import os
import re
import time
import glob

import zipfile
import fitz
import redis
import json5 as json

from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from config import REDIS_HOST, REDIS_PORT, REDIS_PWD, LANGUAGE
from tools.toolbox import (
    CatchException, report_execption, predict_no_ui_but_counting_down
)
from tools.pdfreader import PdfContentReader
from tools.logginger import get_logger
from tools.language_setting import set_language

logger = get_logger()

pool = redis.ConnectionPool(
    host=REDIS_HOST,
    port=REDIS_PORT,
    password=REDIS_PWD,
    decode_responses=True,
    encoding='utf-8'
)
r = redis.Redis(connection_pool=pool)

target_requirement = r.get("target_requirement_content")
language_config = set_language(LANGUAGE)
name_title = language_config["name_title"]
email_title = language_config["email_title"]
is_match_title = language_config["is_match_title"]
reason_title = language_config["reason_title"]
example_name = language_config["example_name"]
example_reason = language_config["example_reason"]
request_1 = language_config["request_1"]
request_2 = language_config["request_2"]
sample_data_format = language_config["sample_data_format"]
file_names = language_config["file_names"]
resume_contents = language_config["resume_contents"]
request_3 = language_config["request_3"]
waiting_gpt_response = language_config["waiting_gpt_response"]
parsing_resume = language_config["parsing_resume"]
no_email = language_config["no_email"]
parsing_failed = language_config["parsing_failed"]
parse_success = language_config["parse_success"]
result_file_name = language_config["result_file_name"]
jod_done = language_config["jod_done"]
parsing_complete = language_config["parsing_complete"]
parsing = language_config["parsing"]
no_file_upload = language_config["no_file_upload"]
no_access = language_config["no_access"]
generated_tip2 = language_config["generated_tip2"]
result_zip_file_name = language_config["result_zip_file_name"]
no_target_files = language_config["no_target_files"]


def save_parse_data_to_excel(file_name, data):
    """
    Save the parsed resume data to an Excel file.

    :param file_name: A string representing the name of the original resume file.
    :param data: A list of dictionaries containing the parsed resume data.
    :return: A string representing the path where the Excel file is saved.
    """
    print(f'\nOriginal data: {data}\n')
    logger.info(f'\nOriginal data: {data}\n')

    # Create a new Excel file
    workbook = Workbook()

    # Add a worksheet named Resume Data
    sheet = workbook.active
    sheet.title = "Resume Parsing Summary"

    # Add data to the worksheet
    header = [name_title, email_title, is_match_title, reason_title]
    sheet.append(header)

    # Sort according to is_match condition
    sorted_data = sorted(data, key=lambda x: x['is_match'])
    for item in sorted_data:
        reason = item.get('reason', '')
        email = item.get('email', '')
        row = [str(item['name']), str(email), str(item['is_match']), str(reason).replace("'", '')]
        sheet.append(row)

    # Loop through all cells and set font background color
    for row in sheet.iter_rows():
        for cell in row:
            # Set font color to red or black based on cell value
            style_dict = {
                'yes': {'font': Font(color='FFFFFF'),
                        'fill': PatternFill(start_color='4bc2c5', end_color='4bc2c5', fill_type='solid')},
                'no': {'font': Font(color='FFFFFF'),
                       'fill': PatternFill(start_color='f6003c', end_color='f6003c', fill_type='solid')},
                'default': {'font': Font(color='000000'),
                            'fill': PatternFill()},
            }

            style = style_dict.get(cell.value, style_dict['default'])
            cell.font = style['font']
            cell.fill = style['fill']

    folder_name = "analysis_reports"
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    # Generate the file name with date and time
    target_file_name = file_name + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '.xlsx'

    # Save and close the Excel file
    workbook.save(filename='analysis_reports/' + target_file_name)
    res = f'{generated_tip2}' + os.path.abspath(
        f'./analysis_reports/{target_file_name}')
    print(res)
    logger.info(res)
    return res


def create_pdf_file_zip(files, zip_filename):
    """
    Create a zip file containing the provided PDF files.
    Args:
        files (list): A list of file paths to the PDF files.
        zip_filename (str): The base name of the resulting zip file.
    """
    if len(files) == 0:
        return no_target_files
    else:
        # Get the current local time and format it as a string
        times_str = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        # Construct the final file name with the provided zip_filename and the timestamp
        file_name = f"analysis_reports/{zip_filename}{times_str}.zip"
        # Open a new zip file with the constructed file name
        with zipfile.ZipFile(file_name, 'w') as zipf:
            for file in files:
                # Preserve the original directory structure in the zip file by using the base name of the file
                zipf.write(file, os.path.basename(file))


def parser_resumes(file_manifest, project_folder, top_p, temperature, chatbot, history, systemPromptTxt):
    """
    Batch-parse resumes and return the result as a list of JSON data for each resume.

    :param project_folder:
    :param file_manifest:
    :param top_p: The cumulative probability to cut off during sampling in GPT.
    :param temperature: The softmax temperature value used to scale the logits before random sampling in GPT.
    :param chatbot: A list of tuples containing the conversation history between the user and the chatbot.
    :param history: A list of strings representing the conversation history between the user and the chatbot.
    :param systemPromptTxt: A string representing the system prompt message to be displayed to the user.
    :return: A tuple containing the chatbot history, user history, and a message indicating the operation status.
    """
    target_file_name = []
    all_parse_data = []

    for index, file_path in enumerate(file_manifest):
        pdf_reader = PdfContentReader()
        file_content = pdf_reader.ocr_pdf_content(file_path)
        file_name = os.path.relpath(file_path, project_folder)
        logger.info(file_name)

        result_example = f'{{"name":{example_name},"email":"xm@gmail.com","is_match":"yes","reason":{example_reason}}}'
        i_say = f'{request_1}{target_requirement}{request_2}{sample_data_format}```{result_example}```，' \
                f'{file_names}{file_name},{resume_contents} ```{file_content}```'
        logger.info(i_say)
        i_say_show_user = f'[{index + 1}/{len(file_manifest)}] {request_3}{os.path.abspath(file_name)}'
        logger.info(i_say_show_user)
        chatbot.append((i_say_show_user, waiting_gpt_response))
        logger.info(chatbot)
        yield chatbot, history, parsing_resume
        msg = parsing_resume
        # ** gpt request **
        gpt_say = yield from predict_no_ui_but_counting_down(i_say, i_say_show_user, chatbot, top_p, temperature,
                                                             history=[])
        logger.info(gpt_say)
        time.sleep(30)
        # Regular expression to match JSON data
        json_pattern = r'\{[^{}]*\}'
        # Extract JSON data from string
        match = re.search(json_pattern, str(gpt_say))
        if match:
            json_data = match.group().replace(example_name, file_name).replace("xm@gmail.com", no_email)
        else:
            json_data = f'{{"name":"{file_name}","email":null,"is_match":"no","reason":{parsing_failed}}}'
        all_parse_data.append(json.loads(json_data))
        show_gpt_say = gpt_say.replace(example_name, file_name).replace("xm@gmail.com", no_email)
        chatbot[-1] = (i_say_show_user, show_gpt_say)
        logger.info(chatbot[-1])
        if json.loads(json_data)["is_match"] == "yes":
            target_file_name.append(file_path)
        # history.append(i_say_show_user)
        # history.append(gpt_say)
        # logger.info(history)
        yield chatbot, history, msg

    i_say = f'{parse_success}\n\n'
    chatbot.append((i_say, waiting_gpt_response))
    logger.info(chatbot)
    yield chatbot, history, parsing_resume
    msg = parsing_resume
    # ** gpt request **
    gpt_say = yield from predict_no_ui_but_counting_down(i_say, i_say, chatbot, top_p, temperature,
                                                         history=history)
    logger.info(gpt_say)

    chatbot[-1] = (i_say, gpt_say)
    # history.append(i_say)
    # history.append(gpt_say)
    yield chatbot, history, msg
    res = save_parse_data_to_excel(result_file_name, all_parse_data)
    chatbot.append((jod_done, res))
    create_pdf_file_zip(target_file_name, result_zip_file_name)
    msg = parsing_complete
    yield chatbot, history, msg


@CatchException
def batch_parser_resumes_plus(txt, top_p, temperature, chatbot, history, systemPromptTxt, WEB_PORT):
    """
    Batch-parse resumes and return the result as a list of JSON data for each resume.

    :param txt: A string representing the path of the project folder or an error message.
    :param top_p: The cumulative probability to cut off during sampling in GPT.
    :param temperature: The softmax temperature value used to scale the logits before random sampling in GPT.
    :param chatbot: A list of tuples containing the conversation history between the user and the chatbot.
    :param history: A list of strings representing the conversation history between the user and the chatbot.
    :param systemPromptTxt: A string representing the system prompt message to be displayed to the user.
    :param WEB_PORT: An integer representing the port number for the web server.
    :return: A tuple containing the chatbot history, user history, and a message indicating the operation status.
    """
    yield chatbot, history, parsing

    # Attempt to import dependencies; if missing, suggest installation method
    try:
        import fitz
    except:
        report_execption(chatbot, history,
                         a=f"parse project: {txt}",
                         b=f"Importing software dependencies failed. Using this module requires additional dependencies, installation method```pip install --upgrade pymupdf```。")
        yield chatbot, history, parsing
        return

    # Clear history to prevent input overflow
    history = []

    # Check input parameters; exit if no input parameter is given
    if os.path.exists(txt):
        project_folder = txt
    else:
        if txt == "":
            txt = no_file_upload
        report_execption(chatbot, history, a=f"{txt}",
                         b=f"{no_access}")
        yield chatbot, history, parsing
        return

    # Search for the list of files to be processed
    # file_manifest = [f for f in glob.glob(f'{project_folder}/**/*.pdf', recursive=True)]

    all_files = glob.glob(f'{project_folder}/**/*.*', recursive=True)

    file_manifest = [f for f in all_files if os.path.splitext(f)[1].lower() == '.pdf']

    not_pdf_files = [f for f in all_files if os.path.splitext(f)[1].lower() != '.pdf']

    # Print non-PDF file information
    if not_pdf_files:
        logger.info("The following files are not PDFs:")
        for not_pdf in not_pdf_files:
            logger.warning(not_pdf)
    else:
        logger.info("All files are PDFs.")

    # If no files are found
    if len(file_manifest) == 0:
        report_execption(chatbot, history, a=f"{txt}", b=f"Could not find any .pdf files: {txt}")
        yield chatbot, history, parsing
        return

    # Begin the task execution
    yield from parser_resumes(file_manifest, project_folder, top_p, temperature, chatbot, history, systemPromptTxt)
